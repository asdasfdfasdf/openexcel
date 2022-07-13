import openpyxl
import utils
import datetime

# 放款路径设置
loan_path = 'E:\\BaiduSyncdisk\\微粒贷统计\\每日数据\\20220713\\放款20220713134235.xlsx'
# 还款路径设置
repayment_path = 'E:\\BaiduSyncdisk\\微粒贷统计\\每日数据\\20220713\\还款20220713145650.xlsx'
# 获取今天的日期
today = datetime.datetime.now().strftime('%Y%m%d')
# 设置保存路径
save_path = 'E:\\BaiduSyncdisk\\微粒贷统计\\每日数据\\' + today + '\\' + today + '还款统计.xlsx'
# 定义开始时间
start_time = datetime.datetime(2022, 5, 20, 0, 0, 0)
# 打开放款表
wb_loan = openpyxl.load_workbook(loan_path,data_only=True)
# 打开放款表的sheet
sheet_loan = wb_loan.worksheets[0]
# 获取放款表借据号这一列的数据存入借据号list
loan_id_list = [c.value for c in sheet_loan['C']][1:]
# 打开还款表
wb_repayment = openpyxl.load_workbook(repayment_path,data_only=True)
# 打开还款表的sheet
sheet_repayment = wb_repayment.worksheets[0]
# 获取还款表的所有数据
repayment_data_list = list([c for c in sheet_repayment.values])[1:]
# print(repayment_data_list)
# 遍历repayment_data，获取还款表的借据号
for repayment_data in repayment_data_list:
    # 如果还款表的借据号在放款表的借据号list中，则获取对应还款表的借据号的行数
    if repayment_data[0] in loan_id_list:
        # 获取还款表的借据号的行数
        repayment_row = loan_id_list.index(repayment_data[0]) + 2
        # 获取还款表的还款日期
        repayment_date = datetime.datetime.strptime(repayment_data[5], '%Y-%m-%d')
        print(repayment_date)
        # 计算还款日期与开始时间的差值
        days = (repayment_date - start_time).days
        # 在放款表中写入还款本金
        sheet_loan.cell(row=repayment_row, column=days).value = utils.change_money(repayment_data[-1])
wb_loan.save(save_path)
# # 获取还款表借据号这一列的数据和还款本金这一列的数据存入借据号list
# repayment_id_list = [c.value for c in sheet_repayment['C'][1:]]
# repayment_principal_list = [utils.change_money(c.value) for c in sheet_repayment['J'][1:]]
# # 获取还款时间这一列的数据存入还款时间list
# repayment_time_list = [utils.change_time(c.value) for c in sheet_repayment['F'][1:]]
# # 将repayment_id_list，repayment_principal_list，repayment_time_list三个list合并成一个list
# repayment_list = list(zip(repayment_id_list, repayment_principal_list, repayment_time_list))
# print(repayment_list)
# # 定义一个空的list，用来存放每一天的每个借据号的还款金额
# repayment_list_every_day = []


# # 判断repayment_list中的借据号是否在loan_id_list中
# for repayment_id in repayment_list:
#     if repayment_id[0] in loan_id_list:
#         # 如果在，写入还款表对应的借据号的还款本金
#         sheet_repayment.cell(row=repayment_id_list.index(repayment_id[0]) + 2, column=11).value = repayment_id[1]
#
